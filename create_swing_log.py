from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"C:\Users\tsnat\tradingview-mcp-jackson\2026-swing-log.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(name="Arial", size=10, color="276221")
RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
RED_FONT   = Font(name="Arial", size=10, color="9C0006")
STD_FONT   = Font(name="Arial", size=10)

COLS = ["Date","Time ET","Symbol","Strategy","Signal","Price","Change %","Session","Result","P&L","Notes"]
WIDTHS = [12, 10, 10, 14, 14, 10, 10, 9, 9, 10, 30]

thin = Side(style="thin", color="CCCCCC")
border = Border(bottom=thin)


def make_month_sheet(wb, month_name):
    ws = wb.create_sheet(month_name)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for i, (col, w) in enumerate(zip(COLS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    return ws


def make_summary_sheet(wb):
    ws = wb.create_sheet("Summary")

    # Title
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "Swing Signal Log — 2026"
    title.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title.fill = HDR_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    hdrs = ["Month","Total Signals","Swing Low","Flip Short","Win","Loss","BE","Win Rate"]
    col_w = [12, 14, 12, 13, 8, 8, 8, 10]
    for i, (h, w) in enumerate(zip(hdrs, col_w), start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.fill = HDR_FILL
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = HDR_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = w

    months = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]

    grey = PatternFill("solid", fgColor="F2F2F2")

    for idx, month in enumerate(months):
        row = idx + 4
        fill = PatternFill("solid", fgColor="FFFFFF") if idx % 2 == 0 else grey

        # Only April has an actual sheet; others hardcoded 0
        if month == "April":
            b = f"=IFERROR(COUNTA(April!E:E)-1,0)"
            c = f"=IFERROR(COUNTIF(April!E:E,\"Swing Low\"),0)"
            d = f"=IFERROR(COUNTIF(April!E:E,\"Flip Short\"),0)"
            e = f"=IFERROR(COUNTIF(April!I:I,\"Win\"),0)"
            f_ = f"=IFERROR(COUNTIF(April!I:I,\"Loss\"),0)"
            g = f"=IFERROR(COUNTIF(April!I:I,\"BE\"),0)"
        else:
            b, c, d, e, f_, g = 0, 0, 0, 0, 0, 0

        e_col = get_column_letter(5)
        f_col = get_column_letter(6)
        g_col = get_column_letter(7)

        values = [month, b, c, d, e, f_, g,
                  f"=IFERROR(IF({e_col}{row}+{f_col}{row}+{g_col}{row}=0,\"-\",{e_col}{row}/({e_col}{row}+{f_col}{row}+{g_col}{row})),\"-\")"]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            if col == 1:
                cell.font = Font(name="Arial", size=10, bold=True)

    # Totals row
    total_row = 16
    ws.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", size=10, bold=True)
    for col in range(2, 8):
        letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col,
                       value=f"=SUM({letter}4:{letter}15)")
        cell.font = Font(name="Arial", size=10, bold=True)
    wr = ws.cell(row=total_row, column=8,
                 value=f"=IFERROR(IF(E{total_row}+F{total_row}+G{total_row}=0,\"-\",E{total_row}/(E{total_row}+F{total_row}+G{total_row})),\"-\")")
    wr.font = Font(name="Arial", size=10, bold=True)

    ws.row_dimensions[3].height = 20
    return ws


wb = Workbook()
wb.remove(wb.active)  # remove default sheet

make_month_sheet(wb, "April")
make_summary_sheet(wb)

wb.save(OUTPUT)
print(f"Created: {OUTPUT}")
